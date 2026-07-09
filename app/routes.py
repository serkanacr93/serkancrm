from flask import render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_user, logout_user, login_required, current_user
from app.models import User, Customer, Deal, DealItem, Production, ProductionItem, Shipment, ShipmentItem, CustomerStatement, Reminder, Product, Task, Commission, Invoice, InvoiceItem, CustomerVisit, DailyReport, Payment, PotentialCustomer, PlacesSearchConfig, PlacesSearchLog
from app.pdf_utils import generate_deal_pdf, generate_statement_pdf
from app import db, places_search
from datetime import datetime, timedelta, date
from functools import wraps
from io import BytesIO
import openpyxl

def admin_required(f):
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            flash('Bu işlem için yönetici yetkisi gereklidir.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def register_routes(app):

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('index'))
        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            remember = request.form.get('remember')
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user, remember=bool(remember))
                user.last_login = datetime.utcnow()
                db.session.commit()
                flash('Hoş geldiniz!', 'success')
                return redirect(request.args.get('next') or url_for('index'))
            flash('Kullanıcı adı veya şifre hatalı.', 'danger')
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        flash('Çıkış yapıldı.', 'info')
        return redirect(url_for('login'))

    @app.route('/users')
    @admin_required
    def users():
        users = User.query.order_by(User.created_at.desc()).all()
        return render_template('users.html', users=users)

    @app.route('/users/add', methods=['GET', 'POST'])
    @admin_required
    def add_user():
        if request.method == 'POST':
            user = User(
                username=request.form['username'],
                email=request.form['email'],
                full_name=request.form.get('full_name'),
                role=request.form.get('role', 'user')
            )
            user.set_password(request.form['password'])
            db.session.add(user)
            db.session.commit()
            flash('Kullanıcı eklendi.', 'success')
            return redirect(url_for('users'))
        return render_template('add_user.html')

    @app.route('/users/<int:id>/edit', methods=['GET', 'POST'])
    @admin_required
    def edit_user(id):
        user = User.query.get_or_404(id)
        if request.method == 'POST':
            user.username = request.form['username']
            user.email = request.form['email']
            user.full_name = request.form.get('full_name')
            user.role = request.form.get('role', 'user')
            user.is_active = 'is_active' in request.form
            if request.form.get('password'):
                user.set_password(request.form['password'])
            db.session.commit()
            flash('Kullanıcı güncellendi.', 'success')
            return redirect(url_for('users'))
        return render_template('edit_user.html', user=user)

    @app.route('/')
    @login_required
    def index():
        today = datetime.now().date()
        
        customers = Customer.query.count()
        deals = Deal.query.count()
        total_value = db.session.query(db.func.sum(Deal.value)).scalar() or 0
        active_deals = Deal.query.filter(Deal.stage.notin_(['kazanilan', 'kaybedilen'])).count()
        production_count = Production.query.count()
        pending_shipments = Shipment.query.filter(Shipment.status.notin_(['teslim_edildi'])).count()
        product_count = Product.query.count()
        low_stock_products = Product.query.filter(Product.stock_quantity <= Product.min_stock).order_by(Product.stock_quantity).all()
        low_stock = len(low_stock_products)

        pending_price_reports = DailyReport.query.filter_by(status='fiyat_verilecek').order_by(DailyReport.report_date.asc()).all()
        pending_price_list = [(r, (today - r.report_date).days) for r in pending_price_reports]

        expiring_deals = Deal.query.filter(
            Deal.valid_until <= today + timedelta(days=2),
            Deal.valid_until >= today,
            Deal.stage.notin_(['kazanilan', 'kaybedilen', 'revize'])
        ).all()
        
        expired_deals = Deal.query.filter(
            Deal.valid_until < today,
            Deal.stage.notin_(['kazanilan', 'kaybedilen', 'revize'])
        ).all()
        
        unread_reminders = Reminder.query.filter_by(is_read=False).count()
        
        recent_deals = Deal.query.order_by(Deal.created_at.desc()).limit(5).all()
        recent_customers = Customer.query.order_by(Customer.created_at.desc()).limit(5).all()
        
        upcoming_tasks = Task.query.filter(
            Task.due_date >= today,
            Task.status.notin_(['tamamlandi'])
        ).order_by(Task.due_date).limit(5).all()
        
        today_tasks = Task.query.filter(
            Task.due_date == today,
            Task.status.notin_(['tamamlandi'])
        ).all()
        
        monthly_sales_rows = db.session.query(
            db.func.to_char(Deal.created_at, 'YYYY-MM').label('month'),
            db.func.sum(Deal.value).label('total')
        ).filter(Deal.stage == 'kazanilan').group_by(db.func.to_char(Deal.created_at, 'YYYY-MM')).order_by(db.text('1 DESC')).limit(6).all()
        monthly_sales = [(r.month, r.total) for r in monthly_sales_rows]

        stage_stats = db.session.query(
            Deal.stage,
            db.func.count(Deal.id)
        ).group_by(Deal.stage).all()
        
        customer_types = db.session.query(
            db.func.count(Customer.id)
        ).filter(Customer.company_name.isnot(None)).scalar() or 0
        individual_customers = customers - customer_types
        
        # Yeni müşteri istatistikleri
        total_visits = CustomerVisit.query.count()
        recent_visits = CustomerVisit.query.order_by(CustomerVisit.visit_date.desc()).limit(5).all()
        
        # Ortalama sipariş dönüşümü (ilk siparişi olan müşteri / toplam müşteri)
        customers_with_orders = db.session.query(Customer.id).join(Deal).filter(Deal.stage == 'kazanilan').distinct().count()
        conversion_rate = (customers_with_orders / customers * 100) if customers > 0 else 0
        
        # Müşteri ortalama yaşı (ilk müşterinin eklenme tarihinden bugüne)
        first_customer = Customer.query.order_by(Customer.created_at.asc()).first()
        avg_customer_days = 0
        if first_customer:
            avg_customer_days = (today - first_customer.created_at.date()).days

        # Müşteri Takip Döngüsü: son siparişten itibaren siparis_dongusu_gun
        # kadar süre sonra yeni bir siparişin beklendiği, ve bu tarihe 30 gün
        # veya daha az kaldığı (ya da geçtiği) müşteriler.
        last_order_subq = db.session.query(
            Deal.customer_id,
            db.func.max(Deal.deal_date).label('last_order_date')
        ).filter(Deal.stage == 'kazanilan').group_by(Deal.customer_id).subquery()

        production_cycle_rows = db.session.query(Customer, last_order_subq.c.last_order_date).join(
            last_order_subq, Customer.id == last_order_subq.c.customer_id
        ).filter(Customer.status != 'musteri_degil').all()

        production_cycle_customers = []
        for cust, last_order_date in production_cycle_rows:
            if not last_order_date:
                continue
            next_expected = last_order_date + timedelta(days=cust.siparis_dongusu_gun)
            days_remaining = (next_expected - today).days
            if days_remaining <= 30:
                production_cycle_customers.append({
                    'customer': cust,
                    'last_order_date': last_order_date,
                    'next_expected': next_expected,
                    'days_remaining': days_remaining,
                })
        production_cycle_customers.sort(key=lambda x: x['days_remaining'])
        production_cycle_customers = production_cycle_customers[:10]

        return render_template('index.html',
                             customers=customers, 
                             deals=deals,
                             total_value=total_value,
                             active_deals=active_deals,
                             production_count=production_count,
                             pending_shipments=pending_shipments,
                             product_count=product_count,
                             low_stock=low_stock,
                             low_stock_products=low_stock_products,
                             pending_price_reports=pending_price_reports,
                             pending_price_list=pending_price_list,
                             expiring_deals=expiring_deals,
                             expired_deals=expired_deals,
                             unread_reminders=unread_reminders,
                             recent_deals=recent_deals,
                             recent_customers=recent_customers,
                             upcoming_tasks=upcoming_tasks,
                             today_tasks=today_tasks,
                             monthly_sales=monthly_sales,
                             stage_stats=stage_stats,
                             customer_types=customer_types,
                             individual_customers=individual_customers,
                             total_visits=total_visits,
                             recent_visits=recent_visits,
                             conversion_rate=conversion_rate,
                             avg_customer_days=avg_customer_days,
                             production_cycle_customers=production_cycle_customers)

    @app.route('/reminders')
    @login_required
    def reminders():
        reminders = Reminder.query.order_by(Reminder.remind_date.desc()).all()
        return render_template('reminders.html', reminders=reminders)

    @app.route('/reminders/<int:id>/read', methods=['POST'])
    @login_required
    def mark_reminder_read(id):
        reminder = Reminder.query.get_or_404(id)
        reminder.is_read = True
        db.session.commit()
        flash('Hatırlatma okundu.', 'success')
        return redirect(url_for('reminders'))

    @app.route('/customers')
    @login_required
    def customers():
        search = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        query = Customer.query.filter(Customer.status != 'musteri_degil')
        if search:
            query = query.filter(
                db.or_(
                    Customer.first_name.ilike(f'%{search}%'),
                    Customer.last_name.ilike(f'%{search}%'),
                    Customer.company_name.ilike(f'%{search}%'),
                    Customer.email.ilike(f'%{search}%'),
                    Customer.tax_id.ilike(f'%{search}%'),
                    Customer.phone.ilike(f'%{search}%')
                )
            )
        pagination = query.order_by(Customer.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
        customers = pagination.items
        not_customer_count = Customer.query.filter_by(status='musteri_degil').count()

        never_transacted_count = Customer.query.filter(
            Customer.status != 'musteri_degil',
            ~Customer.id.in_(db.session.query(Deal.customer_id))
        ).count()

        dormant_cutoff = datetime.utcnow() - timedelta(days=90)
        last_deal_subq2 = db.session.query(
            Deal.customer_id,
            db.func.max(Deal.created_at).label('last_deal_at')
        ).group_by(Deal.customer_id).subquery()
        dormant_count = db.session.query(Customer).join(
            last_deal_subq2, Customer.id == last_deal_subq2.c.customer_id
        ).filter(
            Customer.status != 'musteri_degil',
            last_deal_subq2.c.last_deal_at < dormant_cutoff
        ).count()

        return render_template('customers.html', customers=customers, search=search, pagination=pagination,
                                not_customer_count=not_customer_count, never_transacted_count=never_transacted_count,
                                dormant_count=dormant_count)

    @app.route('/customers/<int:id>/mark-not-customer', methods=['POST'])
    @login_required
    def mark_not_customer(id):
        customer = Customer.query.get_or_404(id)
        customer.status = 'musteri_degil'
        db.session.commit()
        flash(f'"{customer.display_name}" "Müşteri Değil" olarak işaretlendi ve listeden gizlendi.', 'success')
        return redirect(url_for('customers'))

    @app.route('/customers/not-customer')
    @login_required
    def not_customer_list():
        items = Customer.query.filter_by(status='musteri_degil').order_by(Customer.updated_at.desc()).all()
        return render_template('customers_not_customer.html', customers=items)

    @app.route('/customers/not-customer/restore', methods=['POST'])
    @login_required
    def restore_not_customer():
        ids = request.form.getlist('customer_ids')
        if not ids:
            flash('Hiçbir kayıt seçmediniz.', 'warning')
            return redirect(url_for('not_customer_list'))
        customers = Customer.query.filter(Customer.id.in_(ids), Customer.status == 'musteri_degil').all()
        for c in customers:
            c.status = 'aktif'
        db.session.commit()
        flash(f'{len(customers)} kayıt geri alındı, tekrar normal müşteri listesinde görünecek.', 'success')
        return redirect(url_for('not_customer_list'))

    @app.route('/customers/not-customer/delete', methods=['POST'])
    @login_required
    def delete_not_customer():
        ids = request.form.getlist('customer_ids')
        if not ids:
            flash('Hiçbir kayıt seçmediniz.', 'warning')
            return redirect(url_for('not_customer_list'))

        customers = Customer.query.filter(Customer.id.in_(ids), Customer.status == 'musteri_degil').all()
        deleted = 0
        blocked = []
        for c in customers:
            has_deps = (Deal.query.filter_by(customer_id=c.id).first()
                        or Payment.query.filter_by(customer_id=c.id).first()
                        or CustomerStatement.query.filter_by(customer_id=c.id).first())
            if has_deps:
                blocked.append(c.display_name)
                continue
            db.session.delete(c)
            deleted += 1
        db.session.commit()

        msg = f'{deleted} kayıt kalıcı olarak silindi.'
        if blocked:
            msg += f' {len(blocked)} kayıt bağlı teklif/ödeme/ekstre kaydı olduğu için silinemedi: {", ".join(blocked[:5])}'
            if len(blocked) > 5:
                msg += ' ...'
        flash(msg, 'warning' if blocked else 'success')
        return redirect(url_for('not_customer_list'))

    @app.route('/customers/never-transacted')
    @login_required
    def never_transacted_customers():
        page = request.args.get('page', 1, type=int)
        query = Customer.query.filter(
            Customer.status != 'musteri_degil',
            ~Customer.id.in_(db.session.query(Deal.customer_id))
        ).order_by(Customer.created_at.desc())
        pagination = query.paginate(page=page, per_page=50, error_out=False)
        return render_template('customers_never_transacted.html', customers=pagination.items, pagination=pagination)

    @app.route('/customers/dormant')
    @login_required
    def dormant_customers():
        page = request.args.get('page', 1, type=int)
        cutoff = datetime.utcnow() - timedelta(days=90)

        last_deal_subq = db.session.query(
            Deal.customer_id,
            db.func.max(Deal.created_at).label('last_deal_at')
        ).group_by(Deal.customer_id).subquery()

        query = db.session.query(Customer, last_deal_subq.c.last_deal_at).join(
            last_deal_subq, Customer.id == last_deal_subq.c.customer_id
        ).filter(
            Customer.status != 'musteri_degil',
            last_deal_subq.c.last_deal_at < cutoff
        ).order_by(last_deal_subq.c.last_deal_at.asc())

        pagination = query.paginate(page=page, per_page=50, error_out=False)
        rows = [{'customer': c, 'last_deal_at': d, 'days_inactive': (datetime.utcnow() - d).days}
                for c, d in pagination.items]
        return render_template('customers_dormant.html', rows=rows, pagination=pagination)

    @app.route('/customers/add', methods=['GET', 'POST'])
    @login_required
    def add_customer():
        if request.method == 'POST':
            first_name = request.form.get('first_name') or ''
            last_name = request.form.get('last_name') or ''
            
            # Aynı isimden varsa engelle
            existing = Customer.query.filter_by(first_name=first_name, last_name=last_name).first()
            if existing:
                flash(f'"{first_name} {last_name}" adında bir müşteri zaten kayıtlı!', 'danger')
                return render_template('add_customer.html')
            
            customer = Customer(
                first_name=first_name or None,
                last_name=last_name or None,
                email=request.form.get('email') or None,
                phone=request.form.get('phone'),
                company_name=request.form.get('company_name'),
                tax_office=request.form.get('tax_office'),
                tax_id=request.form.get('tax_id'),
                trade_registry=request.form.get('trade_registry'),
                company_phone=request.form.get('company_phone'),
                company_address=request.form.get('company_address'),
                company_email=request.form.get('company_email'),
                company_website=request.form.get('company_website'),
                contact_person=request.form.get('contact_person'),
                contact_title=request.form.get('contact_title'),
                contact_phone=request.form.get('contact_phone'),
                contact_email=request.form.get('contact_email'),
                address=request.form.get('address'),
                notes=request.form.get('notes')
            )
            db.session.add(customer)
            db.session.commit()
            flash('Müşteri eklendi!', 'success')
            return redirect(url_for('customers'))
        return render_template('add_customer.html')

    @app.route('/customers/import', methods=['GET', 'POST'])
    @login_required
    def import_customers():
        if request.method == 'POST':
            file = request.files.get('file')
            if not file or file.filename == '':
                flash('Lütfen bir dosya seçin.', 'danger')
                return render_template('import_customers.html')
            
            added = 0
            skipped = 0
            errors = []
            
            if file.filename.endswith('.csv'):
                import csv
                import io
                stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
                reader = csv.DictReader(stream)
                for i, row in enumerate(reader, 1):
                    try:
                        first_name = (row.get('Ad') or row.get('ad') or row.get('first_name') or '').strip()
                        last_name = (row.get('Soyad') or row.get('soyad') or row.get('last_name') or '').strip()
                        phone = (row.get('Telefon') or row.get('telefon') or row.get('phone') or row.get('GSM') or '').strip()
                        
                        if not first_name and not last_name and not phone:
                            continue
                        
                        existing = Customer.query.filter_by(first_name=first_name, last_name=last_name).first()
                        if existing:
                            skipped += 1
                            continue
                        
                        customer = Customer(
                            first_name=first_name or None,
                            last_name=last_name or None,
                            phone=phone or None,
                            email=(row.get('E-posta') or row.get('email') or row.get('Mail') or '').strip() or None
                        )
                        db.session.add(customer)
                        added += 1
                    except Exception as e:
                        errors.append(f'Satır {i}: {str(e)}')
                
                db.session.commit()
                msg = f'{added} müşteri eklendi.'
                if skipped:
                    msg += f' {skipped} kayıt (aynı isim) atlandı.'
                if errors:
                    msg += f' {len(errors)} hata oluştu.'
                flash(msg, 'success' if added > 0 else 'info')
                if errors:
                    for e in errors[:5]:
                        flash(e, 'warning')
                        
            elif file.filename.endswith(('.xls', '.xlsx')):
                import openpyxl
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        row_dict = dict(zip(headers, [str(v or '') for v in row]))
                        first_name = (row_dict.get('Ad') or row_dict.get('ad') or row_dict.get('first_name') or '').strip()
                        last_name = (row_dict.get('Soyad') or row_dict.get('soyad') or row_dict.get('last_name') or '').strip()
                        phone = (row_dict.get('Telefon') or row_dict.get('telefon') or row_dict.get('phone') or row_dict.get('GSM') or '').strip()
                        
                        if not first_name and not last_name and not phone:
                            continue
                        
                        existing = Customer.query.filter_by(first_name=first_name, last_name=last_name).first()
                        if existing:
                            skipped += 1
                            continue
                        
                        customer = Customer(
                            first_name=first_name or None,
                            last_name=last_name or None,
                            phone=phone or None,
                            email=(row_dict.get('E-posta') or row_dict.get('email') or row_dict.get('Mail') or '').strip() or None
                        )
                        db.session.add(customer)
                        added += 1
                    except Exception as e:
                        errors.append(f'Satır {i}: {str(e)}')
                
                db.session.commit()
                msg = f'{added} müşteri eklendi.'
                if skipped:
                    msg += f' {skipped} kayıt (aynı isim) atlandı.'
                if errors:
                    msg += f' {len(errors)} hata oluştu.'
                flash(msg, 'success' if added > 0 else 'info')
                if errors:
                    for e in errors[:5]:
                        flash(e, 'warning')
            else:
                flash('Yalnızca CSV veya Excel (.xls/.xlsx) dosyaları desteklenir.', 'danger')
                return render_template('import_customers.html')
            
            return redirect(url_for('customers'))
        
        return render_template('import_customers.html')

    @app.route('/customers/<int:id>')
    @login_required
    def customer_detail(id):
        customer = Customer.query.get_or_404(id)
        deals = Deal.query.filter_by(customer_id=id).order_by(Deal.created_at.desc()).all()
        statements = CustomerStatement.query.filter_by(customer_id=id).order_by(CustomerStatement.created_at.desc()).all()
        total_debit = sum(s.amount for s in statements if s.type == 'borc')
        total_credit = sum(s.amount for s in statements if s.type == 'alacak')
        
        # Günlük raporları ekle - önce customer_id ile, sonra isim/telefon ile eşleşenleri bul
        daily_reports = DailyReport.query.filter(
            db.or_(
                DailyReport.customer_id == customer.id,
                DailyReport.customer_name.ilike(f'%{customer.display_name}%'),
                DailyReport.phone == customer.phone
            )
        ).order_by(DailyReport.report_date.desc()).all()

        # Bekleyen iş emirleri: bu müşterinin tekliflerinden doğan, henüz
        # tamamlanmamış üretim kayıtları (Madde 1 - Üretim İş Emri Otomasyonu)
        pending_productions = sorted(
            [d.production for d in deals if d.production and d.production.status != 'tamamlandi'],
            key=lambda p: (p.due_date is None, p.due_date)
        )

        return render_template('customer_detail.html', customer=customer, deals=deals,
                             statements=statements, total_debit=total_debit, total_credit=total_credit,
                             daily_reports=daily_reports, pending_productions=pending_productions)

    @app.route('/customers/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_customer(id):
        customer = Customer.query.get_or_404(id)
        if request.method == 'POST':
            for field in ['first_name', 'last_name', 'email', 'phone', 'company_name', 'tax_office', 'tax_id',
                         'trade_registry', 'company_phone', 'company_address', 'company_email', 'company_website',
                         'contact_person', 'contact_title', 'contact_phone', 'contact_email', 'address', 'notes', 'status']:
                setattr(customer, field, request.form.get(field) or getattr(customer, field))
            siparis_dongusu = request.form.get('siparis_dongusu_gun')
            if siparis_dongusu:
                customer.siparis_dongusu_gun = int(siparis_dongusu)
            db.session.commit()
            flash('Müşteri güncellendi!', 'success')
            return redirect(url_for('customer_detail', id=id))
        return render_template('edit_customer.html', customer=customer)

    @app.route('/customers/merge-by-phone', methods=['POST'])
    @login_required
    def merge_customers_by_phone():
        """Aynı telefon numarasına sahip müşterileri birleştir"""
        phone = request.form.get('phone', '').strip()
        if not phone:
            flash('Telefon numarası gereklidir.', 'danger')
            return redirect(url_for('customers'))
        
        # Aynı telefona sahip tüm müşterileri bul
        customers = Customer.query.filter_by(phone=phone).all()
        
        if len(customers) <= 1:
            flash('Bu telefon numarası ile tekrar eden müşteri bulunamadı.', 'info')
            return redirect(url_for('customers'))
        
        # En eski müşteriyi ana müşteri olarak seç
        main_customer = sorted(customers, key=lambda x: x.id)[0]
        duplicate_customers = [c for c in customers if c.id != main_customer.id]
        
        # Tekrarlayan müşterilerin ilişkilerini ana müşteriye aktar
        for dup in duplicate_customers:
            # Teklifleri aktar
            Deal.query.filter_by(customer_id=dup.id).update({'customer_id': main_customer.id})
            # Ekstreleri aktar
            CustomerStatement.query.filter_by(customer_id=dup.id).update({'customer_id': main_customer.id})
            # Günlük raporları güncelle
            DailyReport.query.filter_by(customer_id=dup.id).update({'customer_id': main_customer.id})
            # Ziyaretleri aktar
            CustomerVisit.query.filter_by(customer_id=dup.id).update({'customer_id': main_customer.id})
            # Görevleri aktar
            Task.query.filter_by(customer_id=dup.id).update({'customer_id': main_customer.id})
            # Hatırlatıcıları aktar
            Reminder.query.filter_by(customer_id=dup.id).update({'customer_id': main_customer.id})
            
            # Tekrarlayan müşteriyi sil
            db.session.delete(dup)
        
        db.session.commit()
        flash(f'{len(duplicate_customers)} tekrar eden müşteri birleştirildi! Ana müşteri: {main_customer.display_name}', 'success')
        return redirect(url_for('customers'))

    @app.route('/customers/<int:id>/delete', methods=['POST'])
    @login_required
    def delete_customer(id):
        customer = Customer.query.get_or_404(id)
        db.session.delete(customer)
        db.session.commit()
        flash('Müşteri silindi!', 'success')
        return redirect(url_for('customers'))

    @app.route('/customers/<int:id>/statement/pdf')
    @login_required
    def customer_statement_pdf(id):
        customer = Customer.query.get_or_404(id)
        statements = CustomerStatement.query.filter_by(customer_id=id).order_by(CustomerStatement.created_at.desc()).all()
        total_debit = sum(s.amount for s in statements if s.type == 'borc')
        total_credit = sum(s.amount for s in statements if s.type == 'alacak')
        pdf = generate_statement_pdf(customer, statements, total_debit, total_credit)
        return send_file(pdf, as_attachment=True, download_name=f'ekstre_{customer.last_name}_{datetime.now().strftime("%Y%m%d")}.pdf')

    @app.route('/customers/export/excel')
    @login_required
    def customers_export_excel():
        customers = Customer.query.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Müşteriler'
        headers = ['Ad', 'Soyad', 'Firma', 'E-posta', 'Telefon', 'Vergi No', 'Yetkili', 'Durum']
        ws.append(headers)
        for c in customers:
            ws.append([c.first_name, c.last_name, c.company_name or '', c.email, c.phone or '', 
                       c.tax_id or '', c.contact_person or '', c.status])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'musteriler_{datetime.now().strftime("%Y%m%d")}.xlsx')

    @app.route('/deals')
    @login_required
    def deals():
        search = request.args.get('search', '')
        stage_filter = request.args.get('stage', '')
        tab = request.args.get('tab', 'aktif')
        page = request.args.get('page', 1, type=int)

        base_query = Deal.query
        if not current_user.is_admin:
            base_query = base_query.filter(Deal.user_id == current_user.id)

        RESOLVED_STAGES = ('kazanilan', 'kaybedilen', 'revize')

        def apply_tab(q, t):
            if t == 'aktif':
                return q.filter(~Deal.stage.in_(RESOLVED_STAGES))
            if t == 'kazanilan':
                return q.filter(Deal.stage == 'kazanilan')
            if t == 'kaybedilen':
                return q.filter(Deal.stage == 'kaybedilen')
            return q  # 'tumu'

        tab_counts = {
            t: apply_tab(base_query, t).count()
            for t in ('aktif', 'kazanilan', 'kaybedilen', 'tumu')
        }

        query = apply_tab(base_query, tab)
        if search:
            query = query.filter(db.or_(
                Deal.title.ilike(f'%{search}%'),
                Customer.first_name.ilike(f'%{search}%'),
                Customer.last_name.ilike(f'%{search}%'),
                Customer.company_name.ilike(f'%{search}%')
            )).join(Customer)
        if stage_filter:
            query = query.filter(Deal.stage == stage_filter)
        pagination = query.order_by(Deal.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
        deals = pagination.items
        return render_template('deals.html', deals=deals, search=search, stage_filter=stage_filter,
                                tab=tab, tab_counts=tab_counts, pagination=pagination)

    @app.route('/deals/add', methods=['GET', 'POST'])
    @login_required
    def add_deal():
        if request.method == 'POST':
            today = datetime.now().date()
            # Otomatik sıralı teklif numarası
            last_deal = Deal.query.order_by(Deal.deal_no.desc()).first()
            next_no = (last_deal.deal_no + 1) if last_deal and last_deal.deal_no else 1
            
            title = request.form.get('title', '').strip()
            if not title:
                # Müşteri adına göre otomatik başlık oluştur: "05.07.2026 - Lema Ambalaj"
                customer = Customer.query.get(int(request.form['customer_id']))
                customer_name = customer.display_name if customer else "Müşteri"
                title = f"{today.strftime('%d.%m.%Y')} - {customer_name}"
            
            deal = Deal(
                deal_no=next_no,
                title=title,
                stage=request.form.get('stage', 'yeni'),
                probability=int(request.form.get('probability', 0)),
                deal_date=today,
                expected_close=datetime.strptime(request.form['expected_close'], '%Y-%m-%d').date() if request.form.get('expected_close') else None,
                valid_until=today + timedelta(days=7),
                vat_rate=float(request.form.get('vat_rate', 20)),
                notes=request.form.get('notes'),
                customer_id=int(request.form['customer_id']),
                user_id=current_user.id
            )
            db.session.add(deal)
            db.session.flush()
            
            i = 0
            while f'desc_{i}' in request.form:
                qty = float(request.form[f'qty_{i}'])
                price = float(request.form[f'price_{i}'])
                item = DealItem(
                    description=request.form[f'desc_{i}'],
                    quantity=qty,
                    unit=request.form.get(f'unit_{i}', 'adet'),
                    unit_price=price,
                    total_price=qty * price,
                    deal_id=deal.id
                )
                db.session.add(item)
                i += 1
            
            db.session.flush()
            deal.calculate_totals()
            
            reminder = Reminder(
                customer_id=deal.customer_id,
                deal_id=deal.id,
                title=f'Teklif Süresi Doluyor: {deal.title}',
                message=f'TKL-{deal.id:05d} teklifinin geçerlilik süresi {deal.valid_until.strftime("%d.%m.%Y")} tarihinde doluyor.',
                remind_date=deal.valid_until - timedelta(days=1)
            )
            db.session.add(reminder)
            db.session.commit()
            flash(f'Teklif oluşturuldu! KDV dahil: {deal.value:,.2f} ₺', 'success')
            return redirect(url_for('deal_detail', id=deal.id))
        customers = Customer.query.order_by(Customer.company_name, Customer.last_name).all()
        return render_template('add_deal.html', customers=customers, today=datetime.now().date(), 
                             expire_date=datetime.now().date() + timedelta(days=7))

    @app.route('/deals/<int:id>')
    @login_required
    def deal_detail(id):
        deal = Deal.query.get_or_404(id)
        if not current_user.is_admin and deal.user_id != current_user.id:
            flash('Bu teklifi görüntüleme yetkiniz yok.', 'danger')
            return redirect(url_for('deals'))
        return render_template('deal_detail.html', deal=deal)

    @app.route('/deals/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_deal(id):
        deal = Deal.query.get_or_404(id)
        if not current_user.is_admin and deal.user_id != current_user.id:
            flash('Bu teklifi düzenleme yetkiniz yok.', 'danger')
            return redirect(url_for('deals'))
        if deal.stage == 'kazanilan':
            flash('Bu teklif onaylanmış ve üretime aktarılmış, kalemleri artık düzenlenemez.', 'danger')
            return redirect(url_for('deal_detail', id=id))
        if request.method == 'POST':
            deal.title = request.form['title']
            deal.stage = request.form['stage']
            deal.probability = int(request.form.get('probability', 0))
            deal.vat_rate = float(request.form.get('vat_rate', 20))
            deal.expected_close = datetime.strptime(request.form['expected_close'], '%Y-%m-%d').date() if request.form.get('expected_close') else None
            deal.valid_until = datetime.strptime(request.form['valid_until'], '%Y-%m-%d').date() if request.form.get('valid_until') else None
            deal.notes = request.form.get('notes')
            
            DealItem.query.filter_by(deal_id=id).delete()
            i = 0
            while f'desc_{i}' in request.form:
                qty = float(request.form[f'qty_{i}'])
                price = float(request.form[f'price_{i}'])
                item = DealItem(description=request.form[f'desc_{i}'], quantity=qty, unit=request.form.get(f'unit_{i}', 'adet'),
                               unit_price=price, total_price=qty * price, deal_id=deal.id)
                db.session.add(item)
                i += 1
            
            db.session.flush()
            deal.calculate_totals()
            db.session.commit()
            flash('Teklif güncellendi!', 'success')
            return redirect(url_for('deal_detail', id=id))
        customers = Customer.query.order_by(Customer.company_name, Customer.last_name).all()
        return render_template('edit_deal.html', deal=deal, customers=customers)

    @app.route('/deals/<int:id>/delete', methods=['POST'])
    @login_required
    def delete_deal(id):
        deal = Deal.query.get_or_404(id)
        if not current_user.is_admin and deal.user_id != current_user.id:
            flash('Bu teklifi silme yetkiniz yok.', 'danger')
            return redirect(url_for('deals'))

        blockers = []
        if deal.production:
            blockers.append('üretim')
        if Commission.query.filter_by(deal_id=id).first():
            blockers.append('prim')
        if deal.invoices:
            blockers.append('fatura')
        if deal.statements:
            blockers.append('cari ekstre')

        if blockers:
            flash(
                'Bu teklife bağlı ' + ', '.join(blockers) + ' kaydı var, '
                'önce onları silin veya bu teklifi silemezsiniz.', 'danger'
            )
            return redirect(url_for('deal_detail', id=id))

        Reminder.query.filter_by(deal_id=id).delete()
        Task.query.filter_by(deal_id=id).delete()
        db.session.delete(deal)
        db.session.commit()
        flash('Teklif silindi!', 'success')
        return redirect(url_for('deals'))

    @app.route('/deals/<int:id>/pdf')
    @login_required
    def deal_pdf(id):
        import os
        deal = Deal.query.get_or_404(id)
        if not current_user.is_admin and deal.user_id != current_user.id:
            flash('Bu teklifin PDF\'ini görüntüleme yetkiniz yok.', 'danger')
            return redirect(url_for('deals'))
        pdf = generate_deal_pdf(deal)
        
        # Teklifler klasörünü oluştur (yoksa)
        teklifler_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Teklifler')
        os.makedirs(teklifler_dir, exist_ok=True)
        
        # Dosya adını teklif başlığından oluştur (Türkçe karakterleri temizle)
        safe_title = deal.title.replace(' ', '_').replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
        filename = f'teklif_{safe_title}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        filepath = os.path.join(teklifler_dir, filename)
        
        # PDF'i dosyaya kaydet
        with open(filepath, 'wb') as f:
            f.write(pdf.read())
        pdf.seek(0)
        
        # Dosyayı kullanıcıya gönder (indir)
        return send_file(filepath, as_attachment=True, download_name=filename)

    @app.route('/deals/<int:id>/revise', methods=['POST'])
    @login_required
    def revise_deal(id):
        deal = Deal.query.get_or_404(id)
        if not current_user.is_admin and deal.user_id != current_user.id:
            flash('Bu teklifi revize etme yetkiniz yok.', 'danger')
            return redirect(url_for('deals'))
        today = datetime.now().date()
        revize_count = Deal.query.filter(Deal.title.like(f'%{deal.title}%')).count()
        new_deal = Deal(
            title=f"{deal.title} (Revize {revize_count})", stage='teklif', probability=deal.probability,
            deal_date=today, expected_close=deal.expected_close, valid_until=today + timedelta(days=7),
            vat_rate=deal.vat_rate, notes=f"Revize. Orijinal: {deal.notes or ''}", customer_id=deal.customer_id,
            user_id=deal.user_id
        )
        db.session.add(new_deal)
        db.session.flush()
        for item in deal.items:
            db.session.add(DealItem(description=item.description, quantity=item.quantity, unit=item.unit,
                                   unit_price=item.unit_price, total_price=item.total_price, deal_id=new_deal.id))
        db.session.flush()
        new_deal.calculate_totals()
        deal.stage = 'revize'
        db.session.commit()
        flash(f'Revize teklif oluşturuldu! KDV dahil: {new_deal.value:,.2f} ₺', 'success')
        return redirect(url_for('deal_detail', id=new_deal.id))

    @app.route('/deals/<int:id>/approve', methods=['GET', 'POST'])
    @login_required
    def approve_deal(id):
        deal = Deal.query.get_or_404(id)
        if not current_user.is_admin and deal.user_id != current_user.id:
            flash('Bu teklifi onaylama yetkiniz yok.', 'danger')
            return redirect(url_for('deals'))

        if request.method == 'POST':
            deal.stage = 'kazanilan'
            deal.user_id = current_user.id
            
            production = Production(deal_id=deal.id, status='beklemede', start_date=datetime.now().date(),
                                     due_date=deal.expected_close)
            db.session.add(production)
            db.session.flush()
            
            # Teklif kalemlerinden ProductionItem'ları oluştur
            for item in deal.items:
                prod_item = ProductionItem(
                    production_id=production.id,
                    deal_item_id=item.id,
                    description=item.description,
                    planned_quantity=item.quantity,
                    produced_quantity=0,
                    unit=item.unit,
                    status='bekleniyor'
                )
                db.session.add(prod_item)
            
            statement = CustomerStatement(customer_id=deal.customer_id, deal_id=deal.id, type='satis', amount=deal.value,
                                         description=f'Satış: {deal.title} (KDV Dahil: {deal.value:,.2f} ₺)')
            db.session.add(statement)
            
            # Manuel prim oranını al
            manual_rate = request.form.get('commission_rate')
            if manual_rate:
                rate = float(manual_rate)
                customer_type = 'manuel'
            else:
                prev_sales = Deal.query.filter(Deal.customer_id == deal.customer_id, Deal.stage == 'kazanilan', Deal.id != deal.id).count()
                is_new = prev_sales == 0
                rate = 1.5 if is_new else 1.0
                customer_type = 'yeni' if is_new else 'eski'
            
            commission_amount = deal.subtotal * (rate / 100)
            
            commission = Commission(
                user_id=current_user.id,
                deal_id=deal.id,
                sale_amount=deal.subtotal,
                rate=rate,
                amount=commission_amount,
                customer_type=customer_type,
                status='odenmedi',
                manual_rate=rate if manual_rate else None
            )
            db.session.add(commission)
            db.session.commit()
            
            flash(f'Teklif onaylandı! Üretime aktarıldı. Prim: %{rate} = {commission_amount:,.2f} ₺', 'success')
            return redirect(url_for('production_detail', id=production.id))
        
        # GET isteği - onay formu göster
        prev_sales = Deal.query.filter(Deal.customer_id == deal.customer_id, Deal.stage == 'kazanilan', Deal.id != deal.id).count()
        is_new = prev_sales == 0
        suggested_rate = 1.5 if is_new else 1.0
        
        return render_template('approve_deal.html', deal=deal, suggested_rate=suggested_rate)

    @app.route('/api/customers/search')
    @login_required
    def search_customers():
        query = request.args.get('q', '')
        if len(query) < 2:
            return jsonify([])
        
        customers = Customer.query.filter(
            db.or_(
                Customer.first_name.ilike(f'%{query}%'),
                Customer.last_name.ilike(f'%{query}%'),
                Customer.company_name.ilike(f'%{query}%'),
                Customer.phone.ilike(f'%{query}%')
            )
        ).limit(10).all()
        
        result = []
        for c in customers:
            result.append({
                'id': c.id,
                'name': c.display_name,
                'phone': c.phone or '',
                'email': c.email or ''
            })
        return jsonify(result)

    @app.route('/deals/export/excel')
    @login_required
    def deals_export_excel():
        deals = Deal.query.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Teklifler'
        headers = ['Teklif No', 'Başlık', 'Müşteri', 'Ara Toplam', 'KDV Oranı', 'KDV', 'Toplam', 'Durum', 'Tarih']
        ws.append(headers)
        for d in deals:
            ws.append([f'TKL-{d.id:05d}', d.title, d.customer.display_name, d.subtotal, f'%{d.vat_rate}', 
                       d.vat_amount, d.value, d.stage, d.deal_date.strftime('%d.%m.%Y') if d.deal_date else ''])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'teklifler_{datetime.now().strftime("%Y%m%d")}.xlsx')

    @app.route('/production')
    @login_required
    def production_list():
        productions = Production.query.order_by(Production.created_at.desc()).all()
        return render_template('production_list.html', productions=productions)

    @app.route('/production/<int:id>')
    @login_required
    def production_detail(id):
        production = Production.query.get_or_404(id)
        shipments = Shipment.query.filter_by(production_id=id).order_by(Shipment.created_at.desc()).all()
        return render_template('production_detail.html', production=production, shipments=shipments,
                                today=datetime.now().date())

    @app.route('/production/<int:id>/update-items', methods=['POST'])
    @login_required
    def update_production_items(id):
        production = Production.query.get_or_404(id)
        
        for item in production.items:
            produced_str = request.form.get(f'produced_{item.id}', '').strip()
            if produced_str == '':
                continue
            try:
                produced = float(produced_str)
            except ValueError:
                continue
            item.produced_quantity = produced
            if item.is_produced:
                item.status = 'uretilen'
            elif item.produced_quantity > 0:
                item.status = 'kismi'
            else:
                item.status = 'bekleniyor'
        
        # Üretim durumunu güncelle
        if production.all_items_produced:
            production.status = 'tamamlandi'
        elif production.produced_items_count > 0:
            production.status = 'uretimde'
        else:
            production.status = 'beklemede'
        
        db.session.commit()
        flash('Üretim miktarları güncellendi!', 'success')
        return redirect(url_for('production_detail', id=id))

    @app.route('/production/<int:id>/create-shipment', methods=['GET', 'POST'])
    @login_required
    def create_shipment_from_production(id):
        production = Production.query.get_or_404(id)
        
        if not production.all_items_produced:
            flash('Tüm ürünler üretilmeden sevkiyat oluşturulamaz!', 'danger')
            return redirect(url_for('production_detail', id=id))
        
        if request.method == 'POST':
            shipment = Shipment(
                production_id=id,
                ship_date=datetime.now().date(),
                status='hazirlaniyor',
                notes=request.form.get('notes', '')
            )
            db.session.add(shipment)
            db.session.flush()
            
            i = 0
            while f'desc_{i}' in request.form:
                qty_str = request.form.get(f'qty_{i}', '').strip()
                if not qty_str:
                    i += 1
                    continue
                qty = float(qty_str)
                if qty > 0:
                    w_str = request.form.get(f'weight_{i}', '').strip()
                    weight_kg = float(w_str) if w_str else None
                    p_str = request.form.get(f'price_{i}', '').strip()
                    price = float(p_str) if p_str else 0
                    item = ShipmentItem(
                        shipment_id=shipment.id,
                        production_item_id=int(request.form.get(f'prod_item_id_{i}', 0)) or None,
                        description=request.form[f'desc_{i}'],
                        quantity=qty,
                        unit=request.form.get(f'unit_{i}', 'adet'),
                        weight_kg=weight_kg,
                        unit_price=price,
                        total_price=qty * price
                    )
                    db.session.add(item)
                i += 1
            
            db.session.commit()
            flash(f'SVN-{shipment.id:05d} sevkiyatı oluşturuldu!', 'success')
            return redirect(url_for('shipment_detail', id=shipment.id))
        
        return render_template('create_shipment_from_production.html', production=production)

    @app.route('/shipments/<int:id>')
    @login_required
    def shipment_detail(id):
        shipment = Shipment.query.get_or_404(id)
        return render_template('shipment_detail.html', shipment=shipment)

    @app.route('/production/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_production(id):
        production = Production.query.get_or_404(id)
        if request.method == 'POST':
            production.status = request.form['status']
            production.start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date() if request.form.get('start_date') else None
            production.end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date() if request.form.get('end_date') else None
            production.due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d').date() if request.form.get('due_date') else None
            production.notes = request.form.get('notes')
            db.session.commit()
            flash('Üretim güncellendi!', 'success')
            return redirect(url_for('production_detail', id=id))
        return render_template('edit_production.html', production=production)

    @app.route('/production/<int:id>/shipment/add', methods=['GET', 'POST'])
    @login_required
    def add_shipment(id):
        production = Production.query.get_or_404(id)
        if request.method == 'POST':
            shipment = Shipment(
                production_id=id, quantity=float(request.form['quantity']),
                unit=request.form.get('unit', 'adet'),
                weight_kg=float(request.form['weight_kg']) if request.form.get('weight_kg') else None,
                ship_date=datetime.strptime(request.form['ship_date'], '%Y-%m-%d').date() if request.form.get('ship_date') else datetime.now().date(),
                tracking_number=request.form.get('tracking_number'), carrier=request.form.get('carrier'),
                status=request.form.get('status', 'hazirlaniyor'), notes=request.form.get('notes')
            )
            db.session.add(shipment)
            db.session.commit()
            flash('Sevkiyat eklendi!', 'success')
            return redirect(url_for('production_detail', id=id))
        return render_template('add_shipment.html', production=production)

    @app.route('/shipments')
    @login_required
    def shipment_list():
        shipments = Shipment.query.order_by(Shipment.created_at.desc()).all()
        return render_template('shipment_list.html', shipments=shipments)

    @app.route('/shipments/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_shipment(id):
        shipment = Shipment.query.get_or_404(id)
        if request.method == 'POST':
            shipment.quantity = float(request.form['quantity'])
            shipment.unit = request.form.get('unit', 'adet')
            shipment.weight_kg = float(request.form['weight_kg']) if request.form.get('weight_kg') else None
            shipment.ship_date = datetime.strptime(request.form['ship_date'], '%Y-%m-%d').date() if request.form.get('ship_date') else None
            shipment.tracking_number = request.form.get('tracking_number')
            shipment.carrier = request.form.get('carrier')
            shipment.status = request.form['status']
            shipment.notes = request.form.get('notes')
            db.session.commit()
            flash('Sevkiyat güncellendi!', 'success')
            return redirect(url_for('production_detail', id=shipment.production_id))
        return render_template('edit_shipment.html', shipment=shipment)

    @app.route('/products')
    @login_required
    def products():
        search = request.args.get('search', '')
        if search:
            products = Product.query.filter(db.or_(
                Product.name.ilike(f'%{search}%'), Product.sku.ilike(f'%{search}%'),
                Product.category.ilike(f'%{search}%')
            )).order_by(Product.name).all()
        else:
            products = Product.query.order_by(Product.name).all()
        return render_template('products.html', products=products, search=search)

    @app.route('/products/add', methods=['GET', 'POST'])
    @login_required
    def add_product():
        if request.method == 'POST':
            product = Product(
                name=request.form['name'], sku=request.form.get('sku'),
                description=request.form.get('description'), unit=request.form.get('unit', 'adet'),
                stock_quantity=float(request.form.get('stock_quantity', 0)),
                min_stock=float(request.form.get('min_stock', 0)),
                cost_price=float(request.form.get('cost_price', 0)),
                sell_price=float(request.form.get('sell_price', 0)),
                category=request.form.get('category')
            )
            db.session.add(product)
            db.session.commit()
            flash('Ürün eklendi!', 'success')
            return redirect(url_for('products'))
        return render_template('add_product.html')

    @app.route('/products/<int:id>')
    @login_required
    def product_detail(id):
        product = Product.query.get_or_404(id)
        return render_template('product_detail.html', product=product)

    @app.route('/products/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_product(id):
        product = Product.query.get_or_404(id)
        if request.method == 'POST':
            product.name = request.form['name']
            product.sku = request.form.get('sku')
            product.description = request.form.get('description')
            product.unit = request.form.get('unit', 'adet')
            product.stock_quantity = float(request.form.get('stock_quantity', 0))
            product.min_stock = float(request.form.get('min_stock', 0))
            product.cost_price = float(request.form.get('cost_price', 0))
            product.sell_price = float(request.form.get('sell_price', 0))
            product.category = request.form.get('category')
            product.status = request.form.get('status', 'aktif')
            db.session.commit()
            flash('Ürün güncellendi!', 'success')
            return redirect(url_for('product_detail', id=id))
        return render_template('edit_product.html', product=product)

    @app.route('/products/<int:id>/delete', methods=['POST'])
    @login_required
    def delete_product(id):
        product = Product.query.get_or_404(id)
        db.session.delete(product)
        db.session.commit()
        flash('Ürün silindi!', 'success')
        return redirect(url_for('products'))

    @app.route('/products/export/excel')
    @login_required
    def products_export_excel():
        products = Product.query.all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ürünler'
        headers = ['SKU', 'Ürün Adı', 'Kategori', 'Birim', 'Stok', 'Min Stok', 'Maliyet', 'Satış Fiyatı', 'Durum']
        ws.append(headers)
        for p in products:
            ws.append([p.sku or '', p.name, p.category or '', p.unit, p.stock_quantity, p.min_stock, 
                       p.cost_price, p.sell_price, 'Düşük Stok' if p.is_low_stock else p.status])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f'urunler_{datetime.now().strftime("%Y%m%d")}.xlsx')

    @app.route('/tasks')
    @login_required
    def tasks():
        status_filter = request.args.get('status', '')
        query = Task.query
        if status_filter:
            query = query.filter(Task.status == status_filter)
        tasks = query.order_by(Task.due_date).all()
        return render_template('tasks.html', tasks=tasks, status_filter=status_filter)

    @app.route('/tasks/add', methods=['GET', 'POST'])
    @login_required
    def add_task():
        if request.method == 'POST':
            task = Task(
                title=request.form['title'], description=request.form.get('description'),
                due_date=datetime.strptime(request.form['due_date'], '%Y-%m-%d').date() if request.form.get('due_date') else None,
                due_time=datetime.strptime(request.form['due_time'], '%H:%M').time() if request.form.get('due_time') else None,
                priority=request.form.get('priority', 'orta'),
                category=request.form.get('category'),
                customer_id=int(request.form['customer_id']) if request.form.get('customer_id') else None,
                deal_id=int(request.form['deal_id']) if request.form.get('deal_id') else None,
                user_id=current_user.id
            )
            db.session.add(task)
            db.session.commit()
            flash('Görev eklendi!', 'success')
            return redirect(url_for('tasks'))
        customers = Customer.query.order_by(Customer.last_name).all()
        deals = Deal.query.order_by(Deal.title).all()
        return render_template('add_task.html', customers=customers, deals=deals)

    @app.route('/tasks/<int:id>/complete', methods=['POST'])
    @login_required
    def complete_task(id):
        task = Task.query.get_or_404(id)
        task.status = 'tamamlandi'
        task.completed_at = datetime.utcnow()
        db.session.commit()
        flash('Görev tamamlandı!', 'success')
        return redirect(url_for('tasks'))

    @app.route('/tasks/<int:id>/delete', methods=['POST'])
    @login_required
    def delete_task(id):
        task = Task.query.get_or_404(id)
        db.session.delete(task)
        db.session.commit()
        flash('Görev silindi!', 'success')
        return redirect(url_for('tasks'))

    @app.route('/calendar')
    @login_required
    def calendar():
        year = int(request.args.get('year', datetime.now().year))
        month = int(request.args.get('month', datetime.now().month))
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)
        
        tasks = Task.query.filter(Task.due_date.between(first_day, last_day)).all()
        deals_expiring = Deal.query.filter(Deal.valid_until.between(first_day, last_day)).all()
        
        cal_tasks = {}
        for t in tasks:
            day = t.due_date.day
            if day not in cal_tasks:
                cal_tasks[day] = []
            cal_tasks[day].append(t)
        
        cal_deals = {}
        for d in deals_expiring:
            day = d.valid_until.day
            if day not in cal_deals:
                cal_deals[day] = []
            cal_deals[day].append(d)
        
        return render_template('calendar.html', year=year, month=month, tasks=cal_tasks, deals=cal_deals, first_day=first_day, last_day=last_day, today=datetime.now().date())

    @app.route('/reports')
    @login_required
    def reports():
        stage_stats = db.session.query(Deal.stage, db.func.count(Deal.id), db.func.sum(Deal.value)).group_by(Deal.stage).all()
        recent_deals = Deal.query.order_by(Deal.created_at.desc()).limit(5).all()
        top_customers = db.session.query(Customer, db.func.count(Deal.id).label('deal_count'),
                                        db.func.sum(Deal.value).label('total_value')).join(Deal).group_by(Customer).order_by(db.text('total_value DESC')).limit(5).all()
        
        production_stats = db.session.query(Production.status, db.func.count(Production.id)).group_by(Production.status).all()
        
        monthly_sales_rows2 = db.session.query(
            db.func.to_char(Deal.created_at, 'YYYY-MM').label('month'),
            db.func.sum(Deal.value).label('total')
        ).filter(Deal.stage == 'kazanilan').group_by(db.func.to_char(Deal.created_at, 'YYYY-MM')).order_by(db.text('1 DESC')).limit(12).all()
        monthly_sales = [(r.month, r.total) for r in monthly_sales_rows2]
        return render_template('reports.html', stage_stats=stage_stats, recent_deals=recent_deals,
                             top_customers=top_customers, production_stats=production_stats, monthly_sales=monthly_sales)

    @app.route('/api/chart-data')
    @login_required
    def chart_data():
        stage_stats = db.session.query(Deal.stage, db.func.count(Deal.id)).group_by(Deal.stage).all()
        monthly_sales = db.session.query(
            db.func.to_char(Deal.created_at, 'YYYY-MM').label('month'),
            db.func.sum(Deal.value).label('total')
        ).filter(Deal.stage == 'kazanilan').group_by(db.func.to_char(Deal.created_at, 'YYYY-MM')).order_by(db.text('1 DESC')).limit(6).all()
        
        return jsonify({
            'stages': [{'stage': s[0], 'count': s[1]} for s in stage_stats],
            'sales': [{'month': s[0], 'total': s[1] or 0} for s in monthly_sales]
        })

    @app.route('/commissions')
    @login_required
    def commissions():
        if current_user.is_admin:
            commissions = Commission.query.order_by(Commission.created_at.desc()).all()
        else:
            commissions = Commission.query.filter_by(user_id=current_user.id).order_by(Commission.created_at.desc()).all()
        
        total_pending = sum(c.amount for c in commissions if c.status == 'odenmedi')
        total_paid = sum(c.amount for c in commissions if c.status == 'odendi')
        
        user_stats = {}
        for c in commissions:
            uid = c.user_id
            if uid not in user_stats:
                user_stats[uid] = {'pending': 0, 'paid': 0, 'count': 0}
            user_stats[uid]['count'] += 1
            if c.status == 'odenmedi':
                user_stats[uid]['pending'] += c.amount
            else:
                user_stats[uid]['paid'] += c.amount
        
        return render_template('commissions.html', commissions=commissions, 
                             total_pending=total_pending, total_paid=total_paid, user_stats=user_stats)

    @app.route('/commissions/<int:id>/pay', methods=['POST'])
    @login_required
    def pay_commission(id):
        commission = Commission.query.get_or_404(id)
        commission.status = 'odendi'
        commission.paid_at = datetime.utcnow()
        db.session.commit()
        flash(f'{commission.amount:,.2f} ₺ prim ödendi olarak işaretlendi.', 'success')
        return redirect(url_for('commissions'))

    @app.route('/commissions/pay-all', methods=['POST'])
    @admin_required
    def pay_all_commissions():
        Commission.query.filter_by(status='odenmedi').update({'status': 'odendi', 'paid_at': datetime.utcnow()})
        db.session.commit()
        flash('Tüm ödenmemiş primler ödendi olarak işaretlendi.', 'success')
        return redirect(url_for('commissions'))

    @app.route('/settings')
    @login_required
    def settings():
        db_size = db.session.execute(db.text("SELECT pg_database_size(current_database())")).scalar()

        customer_count = Customer.query.count()
        deal_count = Deal.query.count()
        places_config = places_search.get_config()

        return render_template('settings.html', db_size=db_size, customer_count=customer_count,
                                deal_count=deal_count, places_config=places_config)

    @app.route('/settings/places-toggle', methods=['POST'])
    @login_required
    def toggle_places_search():
        config = places_search.get_config()
        config.enabled = not config.enabled
        db.session.commit()
        flash(f"Otomatik müşteri arama {'aktif' if config.enabled else 'pasif'} edildi.", 'success')
        return redirect(url_for('settings'))

    @app.route('/users/<int:id>/commissions')
    @login_required
    def user_commissions(id):
        user = User.query.get_or_404(id)
        if not current_user.is_admin and current_user.id != id:
            flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
            return redirect(url_for('commissions'))
        
        commissions = Commission.query.filter_by(user_id=id).order_by(Commission.created_at.desc()).all()
        total_pending = sum(c.amount for c in commissions if c.status == 'odenmedi')
        total_paid = sum(c.amount for c in commissions if c.status == 'odendi')
        
        return render_template('user_commissions.html', user=user, commissions=commissions,
                             total_pending=total_pending, total_paid=total_paid)

    # Fatura / İrsaliye Routes
    @app.route('/invoices')
    @login_required
    def invoices():
        search = request.args.get('search', '')
        type_filter = request.args.get('type', '')
        page = request.args.get('page', 1, type=int)
        query = Invoice.query
        if search:
            query = query.join(Customer).filter(db.or_(
                Customer.first_name.ilike(f'%{search}%'),
                Customer.last_name.ilike(f'%{search}%'),
                Customer.company_name.ilike(f'%{search}%')
            ))
        if type_filter:
            query = query.filter(Invoice.type == type_filter)
        pagination = query.order_by(Invoice.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
        invoices = pagination.items
        return render_template('invoices.html', invoices=invoices, search=search, type_filter=type_filter, pagination=pagination)

    @app.route('/invoices/<int:id>')
    @login_required
    def invoice_detail(id):
        invoice = Invoice.query.get_or_404(id)
        return render_template('invoice_detail.html', invoice=invoice)

    @app.route('/deals/<int:id>/create-invoice', methods=['GET', 'POST'])
    @login_required
    def create_invoice_from_deal(id):
        deal = Deal.query.get_or_404(id)
        if not current_user.is_admin and deal.user_id != current_user.id:
            flash('Bu teklif için fatura oluşturma yetkiniz yok.', 'danger')
            return redirect(url_for('deals'))

        if request.method == 'POST':
            inv_type = request.form.get('type', 'fatura')
            
            # Otomatik sıralı fatura/irsaliye no
            last_inv = Invoice.query.order_by(Invoice.invoice_no.desc()).first()
            next_no = (last_inv.invoice_no + 1) if last_inv and last_inv.invoice_no else 1
            
            invoice = Invoice(
                invoice_no=next_no,
                type=inv_type,
                deal_id=deal.id,
                customer_id=deal.customer_id,
                date=datetime.now().date(),
                vat_rate=deal.vat_rate,
                notes=request.form.get('notes', '')
            )
            db.session.add(invoice)
            db.session.flush()
            
            # Teklif kalemlerini kopyala, manuel girilen kg/adet ile güncelle
            i = 0
            while f'desc_{i}' in request.form:
                qty = float(request.form.get(f'qty_{i}', 0))
                price = float(request.form.get(f'price_{i}', 0))
                item = InvoiceItem(
                    invoice_id=invoice.id,
                    description=request.form[f'desc_{i}'],
                    quantity=qty,
                    unit=request.form.get(f'unit_{i}', 'adet'),
                    unit_price=price,
                    total_price=qty * price
                )
                db.session.add(item)
                i += 1
            
            db.session.flush()
            invoice.calculate_totals()
            db.session.commit()
            
            flash(f'{invoice.display_no} - {inv_type} başarıyla oluşturuldu!', 'success')
            return redirect(url_for('invoice_detail', id=invoice.id))
        
        return render_template('create_invoice_from_deal.html', deal=deal)

    @app.route('/visits')
    @login_required
    def visits():
        visits = CustomerVisit.query.order_by(CustomerVisit.visit_date.desc()).all()
        return render_template('visits.html', visits=visits)

    @app.route('/visits/add', methods=['GET', 'POST'])
    @login_required
    def add_visit():
        if request.method == 'POST':
            customer_id = int(request.form['customer_id'])
            visit = CustomerVisit(
                customer_id=customer_id,
                user_id=current_user.id,
                visit_date=datetime.strptime(request.form['visit_date'], '%Y-%m-%d').date() if request.form.get('visit_date') else datetime.now().date(),
                notes=request.form.get('notes'),
                visit_type=request.form.get('visit_type', 'ziyaret')
            )
            db.session.add(visit)
            db.session.commit()
            flash('Ziyaret kaydedildi!', 'success')
            return redirect(url_for('visits'))
        
        # Telefon ile hızlı müşteri bulma
        phone = request.args.get('phone', '')
        customer = None
        if phone:
            customer = Customer.query.filter_by(phone=phone).first()
        
        customers = Customer.query.order_by(Customer.company_name, Customer.last_name).all()
        return render_template('add_visit.html', customers=customers, phone=phone, customer=customer)
    
    @app.route('/api/customers/by-phone')
    @login_required
    def get_customer_by_phone():
        phone = request.args.get('phone', '')
        customer = Customer.query.filter_by(phone=phone).first()
        if customer:
            return jsonify({
                'id': customer.id,
                'name': customer.display_name,
                'phone': customer.phone or '',
                'email': customer.email or ''
            })
        return jsonify(None)

    @app.route('/customers/import/vcf', methods=['GET', 'POST'])
    @login_required
    def import_customers_vcf():
        if request.method == 'POST':
            if 'file' not in request.files:
                flash('Dosya seçilmedi.', 'danger')
                return redirect(url_for('customers'))
            file = request.files['file']
            if file.filename == '':
                flash('Dosya seçilmedi.', 'danger')
                return redirect(url_for('customers'))
            
            try:
                content = file.read().decode('utf-8', errors='ignore')
                lines = content.split('\n')
                
                added = 0
                skipped = 0
                errors = []
                
                count = 0
                current_customer = {}
                
                for line in lines:
                    line = line.strip()
                    if line.startswith('BEGIN:VCARD'):
                        current_customer = {}
                    elif line.startswith('END:VCARD'):
                        if current_customer.get('name') or current_customer.get('phone'):
                            name_parts = current_customer.get('name', '').split()
                            first_name = name_parts[0] if name_parts else None
                            last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else None
                            
                            # Aynı isimde müşteri var mı kontrol et
                            existing = Customer.query.filter_by(first_name=first_name, last_name=last_name).first()
                            if existing:
                                skipped += 1
                                current_customer = {}
                                continue
                            
                            customer = Customer(
                                first_name=first_name,
                                last_name=last_name,
                                phone=current_customer.get('phone'),
                                email=current_customer.get('email'),
                                company_name=current_customer.get('org')
                            )
                            db.session.add(customer)
                            added += 1
                        current_customer = {}
                    elif line.startswith('FN:'):
                        current_customer['name'] = line[3:].strip()
                    elif line.startswith('N:'):
                        parts = line[2:].strip().split(';')
                        if len(parts) >= 2:
                            current_customer['last_name'] = parts[0].strip()
                            current_customer['first_name'] = parts[1].strip()
                    elif 'TEL' in line and ':' in line:
                        phone = line.split(':', 1)[1].strip()
                        if phone:
                            current_customer['phone'] = phone
                    elif 'EMAIL' in line and ':' in line:
                        email = line.split(':', 1)[1].strip()
                        if email:
                            current_customer['email'] = email
                    elif 'ORG' in line and ':' in line:
                        org = line.split(':', 1)[1].strip()
                        if org:
                            current_customer['org'] = org
                
                db.session.commit()
                msg = f'{added} kişi eklendi.'
                if skipped:
                    msg += f' {skipped} kişi (aynı isim) atlandı.'
                flash(msg, 'success' if added > 0 else 'info')
            except Exception as e:
                flash(f'VCF dosyası okuma hatası: {str(e)}', 'danger')
            
            return redirect(url_for('customers'))
        return render_template('import_vcf.html')

    # Günlük Rapor Routes
    @app.route('/daily-reports')
    @login_required
    def daily_reports():
        reports = DailyReport.query.order_by(DailyReport.report_date.desc(), DailyReport.created_at.desc()).all()

        dates = sorted({r.report_date for r in reports}, reverse=True)
        day_cards = []
        for d in dates:
            day_reports = [r for r in reports if r.report_date == d]

            deal_count = Deal.query.filter(Deal.deal_date == d).count()
            avg_deal_value = db.session.query(db.func.avg(Deal.value)).filter(Deal.deal_date == d).scalar() or 0
            shipment_count = Shipment.query.filter(Shipment.ship_date == d).count()

            payments_day = Payment.query.filter(Payment.payment_date == d).all()
            payment_count = len(payments_day)
            payment_total = sum(p.amount for p in payments_day)

            pending_price_count = sum(1 for r in day_reports if r.status == 'fiyat_verilecek')

            top_customer_row = db.session.query(
                Customer.first_name, Customer.last_name, Customer.company_name,
                db.func.count(Deal.id).label('cnt')
            ).join(Deal, Deal.customer_id == Customer.id).filter(Deal.deal_date == d) \
             .group_by(Customer.id, Customer.first_name, Customer.last_name, Customer.company_name) \
             .order_by(db.text('cnt DESC')).first()
            top_customer = None
            if top_customer_row:
                top_customer = top_customer_row.company_name or f"{top_customer_row.first_name} {top_customer_row.last_name}"

            day_cards.append({
                'date': d,
                'reports': day_reports,
                'deal_count': deal_count,
                'avg_deal_value': avg_deal_value,
                'shipment_count': shipment_count,
                'payment_count': payment_count,
                'payment_total': payment_total,
                'pending_price_count': pending_price_count,
                'top_customer': top_customer,
            })

        return render_template('daily_reports.html', day_cards=day_cards)

    @app.route('/daily-reports/add', methods=['GET', 'POST'])
    @login_required
    def add_daily_report():
        today = datetime.now().date()
        
        if request.method == 'POST':
            report_date_str = request.form.get('report_date', '')
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date() if report_date_str else today
            
            customer_names = request.form.getlist('customer_name[]')
            phones = request.form.getlist('phone[]')
            notes_list = request.form.getlist('notes[]')
            
            if not customer_names or not any(customer_names):
                flash('En az bir müşteri adı gereklidir.', 'danger')
                return render_template('add_daily_report.html', today=today)
            
            reports_added = 0
            
            for i, customer_name in enumerate(customer_names):
                customer_name = customer_name.strip()
                if not customer_name:
                    continue
                
                phone = phones[i].strip() if i < len(phones) else ''
                notes = notes_list[i].strip() if i < len(notes_list) else ''
                statuses = request.form.getlist('status[]')
                status = statuses[i].strip() if i < len(statuses) and statuses[i] else 'takip_edilecek'

                # Önce formdan gelen customer_id'yi kontrol et
                customer_ids = request.form.getlist('customer_id[]')
                customer_id = int(customer_ids[i]) if i < len(customer_ids) and customer_ids[i] else None
                
                # Eğer customer_id yoksa, telefon ile müşteri bul
                if not customer_id and phone:
                    matched_customer = Customer.query.filter_by(phone=phone).first()
                    if matched_customer:
                        customer_id = matched_customer.id
                
                # Eğer telefon numarası farklı bir müşteriye aitse, yeni müşteri kaydet
                if phone and customer_id:
                    matched_customer = Customer.query.get(customer_id)
                    if matched_customer:
                        existing_name = matched_customer.display_name
                        if customer_name.lower() != existing_name.lower():
                            # Farklı müşteri - yeni kayıt oluştur
                            name_parts = customer_name.split()
                            first_name = name_parts[0] if name_parts else customer_name
                            last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                            
                            new_customer = Customer(
                                first_name=first_name,
                                last_name=last_name or None,
                                phone=phone
                            )
                            db.session.add(new_customer)
                            db.session.flush()
                            customer_id = new_customer.id
                            flash(f'Yeni müşteri kaydedildi: {customer_name} (Telefon: {phone})', 'success')
                
                report = DailyReport(
                    report_date=report_date,
                    customer_name=customer_name,
                    phone=phone or None,
                    notes=notes,
                    status=status,
                    user_id=current_user.id,
                    customer_id=customer_id
                )
                db.session.add(report)
                reports_added += 1
            
            db.session.commit()
            flash(f'{reports_added} adet günlük rapor eklendi!', 'success')
            return redirect(url_for('daily_reports'))
        
        return render_template('add_daily_report.html', today=today)

    @app.route('/api/customers/search-by-name')
    @login_required
    def search_customers_by_name():
        query = request.args.get('q', '')
        if len(query) < 2:
            return jsonify([])
        
        # Müşteri ID'si, isim, soyisim, firma adı ve telefon ile arama
        customers = Customer.query.filter(
            db.or_(
                Customer.id == query if query.isdigit() else False,
                Customer.first_name.ilike(f'%{query}%'),
                Customer.last_name.ilike(f'%{query}%'),
                Customer.company_name.ilike(f'%{query}%'),
                Customer.phone.ilike(f'%{query}%')
            )
        ).limit(20).all()
        
        result = []
        for c in customers:
            result.append({
                'id': c.id,
                'name': c.display_name,
                'phone': c.phone or '',
                'email': c.email or '',
                'company_name': c.company_name or ''
            })
        return jsonify(result)

    @app.route('/daily-reports/by-date/<date_str>')
    @login_required
    def daily_reports_by_date(date_str):
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = datetime.now().date()
        return redirect(url_for('daily_reports') + f'#gun-{target_date.strftime("%Y%m%d")}')

    @app.route('/daily-reports/<int:id>/delete', methods=['POST'])
    @login_required
    def delete_daily_report(id):
        report = DailyReport.query.get_or_404(id)
        db.session.delete(report)
        db.session.commit()
        flash('Günlük rapor silindi!', 'success')
        return redirect(url_for('daily_reports'))

    # Ödeme Routes
    @app.route('/payments')
    @login_required
    def payments():
        customer_search = request.args.get('customer_search', '')
        status_filter = request.args.get('status', '')
        
        query = Payment.query
        
        if customer_search:
            query = query.join(Customer).filter(
                db.or_(
                    Customer.first_name.ilike(f'%{customer_search}%'),
                    Customer.last_name.ilike(f'%{customer_search}%'),
                    Customer.company_name.ilike(f'%{customer_search}%')
                )
            )
        
        if status_filter:
            query = query.filter(Payment.status == status_filter)
        
        payments = query.order_by(Payment.payment_date.desc()).all()
        return render_template('payments.html', payments=payments, 
                             customer_search=customer_search, status_filter=status_filter)

    @app.route('/payments/add', methods=['GET', 'POST'])
    @login_required
    def add_payment():
        if request.method == 'POST':
            payment = Payment(
                customer_id=int(request.form['customer_id']),
                invoice_id=int(request.form['invoice_id']) if request.form.get('invoice_id') else None,
                amount=float(request.form['amount']),
                payment_date=datetime.strptime(request.form['payment_date'], '%Y-%m-%d').date() if request.form.get('payment_date') else datetime.now().date(),
                payment_method=request.form.get('payment_method'),
                reference_no=request.form.get('reference_no'),
                notes=request.form.get('notes'),
                status=request.form.get('status', 'odendi'),
                user_id=current_user.id
            )
            db.session.add(payment)
            db.session.commit()
            flash('Ödeme kaydedildi!', 'success')
            return redirect(url_for('payments'))
        
        customers = Customer.query.order_by(Customer.company_name, Customer.last_name).all()
        invoices = Invoice.query.order_by(Invoice.date.desc()).all()
        return render_template('add_payment.html', customers=customers, invoices=invoices, today=datetime.now().date())

    @app.route('/payments/<int:id>/delete', methods=['POST'])
    @login_required
    def delete_payment(id):
        payment = Payment.query.get_or_404(id)
        db.session.delete(payment)
        db.session.commit()
        flash('Ödeme kaydı silindi!', 'success')
        return redirect(url_for('payments'))

    @app.route('/api/customer-payments/<int:customer_id>')
    @login_required
    def get_customer_payments(customer_id):
        payments = Payment.query.filter_by(customer_id=customer_id).order_by(Payment.payment_date.desc()).all()
        result = []
        for p in payments:
            result.append({
                'id': p.id,
                'amount': p.amount,
                'date': p.payment_date.strftime('%d.%m.%Y'),
                'method': p.payment_method or '-',
                'status': p.status,
                'invoice_no': p.invoice.display_no if p.invoice else '-'
            })
        return jsonify(result)

    @app.route('/potential-customers')
    @login_required
    def potential_customers():
        search = request.args.get('search', '')
        status_filter = request.args.get('status', '')
        city_filter = request.args.get('city', '')
        sector_filter = request.args.get('sector', '')
        product_filter = request.args.get('product', '')
        page = request.args.get('page', 1, type=int)

        query = PotentialCustomer.query
        if search:
            query = query.filter(db.or_(
                PotentialCustomer.company_name.ilike(f'%{search}%'),
                PotentialCustomer.phone.ilike(f'%{search}%')
            ))
        if status_filter:
            query = query.filter(PotentialCustomer.status == status_filter)
        if city_filter:
            query = query.filter(PotentialCustomer.city == city_filter)
        if sector_filter:
            query = query.filter(PotentialCustomer.sector == sector_filter)
        if product_filter:
            query = query.filter(PotentialCustomer.interested_products.ilike(f'%{product_filter}%'))

        pagination = query.order_by(PotentialCustomer.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
        potentials = pagination.items

        cities = [c[0] for c in db.session.query(PotentialCustomer.city).filter(
            PotentialCustomer.city.isnot(None), PotentialCustomer.city != ''
        ).distinct().order_by(PotentialCustomer.city).all()]

        places_config = places_search.get_config()
        places_stats = {
            'status': places_search.get_status(places_config),
            'today_used': places_search.todays_request_count(),
            'today_limit': places_search.DAILY_REQUEST_LIMIT,
            'today_new': places_search.todays_new_companies(),
            'month': places_search.month_stats(),
            'last_90_days': places_search.last_90_days_stats(),
            'recent_logs': PlacesSearchLog.query.order_by(PlacesSearchLog.run_at.desc()).limit(5).all(),
        }

        return render_template('potential_customers.html', potentials=potentials, pagination=pagination,
                                search=search, status_filter=status_filter, city_filter=city_filter,
                                sector_filter=sector_filter, product_filter=product_filter, cities=cities,
                                sectors=PotentialCustomer.SECTORS, products=PotentialCustomer.PRODUCTS,
                                statuses=PotentialCustomer.STATUSES, places_stats=places_stats,
                                all_cities=places_search.ALL_CITIES, search_sectors=places_search.SEARCH_SECTORS)

    @app.route('/potential-customers/export/excel')
    @login_required
    def potential_customers_export_excel():
        from openpyxl.styles import Font, PatternFill, Alignment

        search = request.args.get('search', '')
        status_filter = request.args.get('status', '')
        city_filter = request.args.get('city', '')
        sector_filter = request.args.get('sector', '')
        product_filter = request.args.get('product', '')

        query = PotentialCustomer.query
        if search:
            query = query.filter(db.or_(
                PotentialCustomer.company_name.ilike(f'%{search}%'),
                PotentialCustomer.phone.ilike(f'%{search}%')
            ))
        if status_filter:
            query = query.filter(PotentialCustomer.status == status_filter)
        if city_filter:
            query = query.filter(PotentialCustomer.city == city_filter)
        if sector_filter:
            query = query.filter(PotentialCustomer.sector == sector_filter)
        if product_filter:
            query = query.filter(PotentialCustomer.interested_products.ilike(f'%{product_filter}%'))
        items = query.order_by(PotentialCustomer.created_at.desc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Potansiyel Müşteriler'
        headers = ['Firma Adı', 'Telefon', 'Adres', 'Şehir', 'Sektör', 'İlgilenilen Ürünler',
                   'Durum', 'Web Sitesi', 'Kaynak', 'Eklenme Tarihi']
        ws.append(headers)
        header_fill = PatternFill(start_color='1a252f', end_color='1a252f', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for pc in items:
            website = ''
            if pc.notes and 'Website:' in pc.notes:
                website = pc.notes.split('Website:', 1)[1].strip()
                if website == '-':
                    website = ''
            ws.append([
                pc.company_name, pc.phone or '', pc.address or '', pc.city or '', pc.sector or '',
                ', '.join(pc.product_list), pc.status or '', website, pc.source or '',
                pc.created_at.strftime('%d.%m.%Y') if pc.created_at else ''
            ])

        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 50)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True,
                          download_name=f'potansiyel_musteriler_{datetime.now().strftime("%Y%m%d")}.xlsx')

    @app.route('/potential-customers/search-now', methods=['POST'])
    @login_required
    def search_places_now():
        selected_cities = request.form.getlist('cities')
        selected_sectors = request.form.getlist('sectors')
        result = places_search.run_batch_search(selected_cities, selected_sectors, triggered_by='manuel')

        if result.get('skipped'):
            flash(result.get('reason', 'Arama yapılamadı.'), 'warning')
        else:
            msg = (f"Arama tamamlandı: {result['combos_run']} kombinasyon tarandı, "
                   f"{result['total_results']} sonuç bulundu, {result['total_new']} yeni firma eklendi "
                   f"({result['total_requests']} istek kullanıldı).")
            if result['combos_skipped']:
                msg += f" {result['combos_skipped']} kombinasyon günlük kota nedeniyle atlandı."
            if result['errors']:
                msg += f" {len(result['errors'])} kombinasyonda hata oluştu."
            flash(msg, 'warning' if result['errors'] or result['combos_skipped'] else 'success')
        return redirect(url_for('potential_customers'))

    @app.route('/potential-customers/add', methods=['GET', 'POST'])
    @login_required
    def add_potential_customer():
        if request.method == 'POST':
            pc = PotentialCustomer(
                company_name=request.form.get('company_name', '').strip(),
                phone=request.form.get('phone'),
                address=request.form.get('address'),
                city=request.form.get('city'),
                sector=request.form.get('sector'),
                interested_products=','.join(request.form.getlist('products')),
                source=request.form.get('source', 'Elle'),
                status=request.form.get('status', 'Aranacak'),
                notes=request.form.get('notes'),
            )
            db.session.add(pc)
            db.session.commit()
            flash('Potansiyel müşteri eklendi!', 'success')
            return redirect(url_for('potential_customers'))
        return render_template('add_potential_customer.html', potential=None,
                                sectors=PotentialCustomer.SECTORS, products=PotentialCustomer.PRODUCTS,
                                sources=PotentialCustomer.SOURCES, statuses=PotentialCustomer.STATUSES)

    @app.route('/potential-customers/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    def edit_potential_customer(id):
        pc = PotentialCustomer.query.get_or_404(id)
        if request.method == 'POST':
            pc.company_name = request.form.get('company_name', '').strip()
            pc.phone = request.form.get('phone')
            pc.address = request.form.get('address')
            pc.city = request.form.get('city')
            pc.sector = request.form.get('sector')
            pc.interested_products = ','.join(request.form.getlist('products'))
            pc.source = request.form.get('source', 'Elle')
            pc.status = request.form.get('status', 'Aranacak')
            pc.notes = request.form.get('notes')
            db.session.commit()
            flash('Potansiyel müşteri güncellendi!', 'success')
            return redirect(url_for('potential_customers'))
        return render_template('add_potential_customer.html', potential=pc,
                                sectors=PotentialCustomer.SECTORS, products=PotentialCustomer.PRODUCTS,
                                sources=PotentialCustomer.SOURCES, statuses=PotentialCustomer.STATUSES)

    @app.route('/potential-customers/<int:id>/delete', methods=['POST'])
    @login_required
    def delete_potential_customer(id):
        pc = PotentialCustomer.query.get_or_404(id)
        db.session.delete(pc)
        db.session.commit()
        flash('Potansiyel müşteri silindi!', 'success')
        return redirect(url_for('potential_customers'))

    @app.route('/potential-customers/<int:id>/convert', methods=['POST'])
    @login_required
    def convert_potential_customer(id):
        pc = PotentialCustomer.query.get_or_404(id)
        if pc.converted_customer_id:
            flash('Bu potansiyel müşteri zaten dönüştürülmüş.', 'warning')
            return redirect(url_for('potential_customers'))

        customer = Customer(
            company_name=pc.company_name,
            phone=pc.phone,
            address=pc.address,
            notes=f'Potansiyel müşteriden dönüştürüldü. {pc.notes or ""}'.strip(),
            status='aktif',
        )
        db.session.add(customer)
        db.session.flush()

        pc.converted_customer_id = customer.id
        pc.status = 'Müşteriye Dönüştürüldü'
        db.session.commit()
        flash(f'"{pc.company_name}" müşteriye dönüştürüldü!', 'success')
        return redirect(url_for('potential_customers'))
